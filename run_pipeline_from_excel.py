"""
run_pipeline_from_excel.py
--------------------------
Run the full attachment pipeline (stages 1–8: INGESTION → COMPLETE) for a
list of purchase requisition numbers read from an Excel file.

Each PR is processed from scratch — any existing pipeline data is wiped and
rebuilt (same behaviour as: python run_pipeline.py --pr-no <PR>).

Usage
-----
    # Column auto-detected (first column, or any column named PURCHASE_REQ_NO):
    python run_pipeline_from_excel.py --excel path/to/ras_list.xlsx

    # Explicit column name:
    python run_pipeline_from_excel.py --excel path/to/ras_list.xlsx --column "PR Number"

    # Specific sheet (default: first sheet):
    python run_pipeline_from_excel.py --excel path/to/ras_list.xlsx --sheet "Sheet2"

Logs
----
    Console : INFO and above
    File    : logs/pipeline_excel_YYYY-MM-DD.log  (DEBUG+, rotated daily, 30 days)
"""

from __future__ import annotations

import argparse
import sys
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

from loguru import logger

from utils.config import AppConfig
from db.connection import connect_with_retry
from db.tables import AzureTables
from pipeline.orchestrator import PipelineOrchestrator

_LOG_DIR = Path("logs")

# Column names that are recognised automatically (case-insensitive).
_AUTO_DETECT_COLUMNS = {"purchase_req_no", "purchase req no", "pr_no", "pr no", "ras_no", "ras no"}


# ── Logging ────────────────────────────────────────────────────────────────────

def _configure_logging() -> None:
    _LOG_DIR.mkdir(exist_ok=True)
    logger.remove()

    logger.add(
        sys.stderr,
        level="INFO",
        format=(
            "<green>{time:YYYY-MM-DD HH:mm:ss}</green> | "
            "<level>{level: <8}</level> | "
            "<cyan>{extra[component]}</cyan> | "
            "{message}"
        ),
        colorize=True,
        backtrace=False,
        diagnose=False,
        filter=lambda r: "component" in r["extra"],
    )
    logger.add(
        sys.stderr,
        level="INFO",
        format=(
            "<green>{time:YYYY-MM-DD HH:mm:ss}</green> | "
            "<level>{level: <8}</level> | "
            "{message}"
        ),
        colorize=True,
        backtrace=False,
        diagnose=False,
        filter=lambda r: "component" not in r["extra"],
    )
    logger.add(
        _LOG_DIR / "pipeline_excel_{time:YYYY-MM-DD}.log",
        level="DEBUG",
        format=(
            "{time:YYYY-MM-DD HH:mm:ss.SSS} | "
            "{level: <8} | "
            "{extra} | "
            "{name}:{function}:{line} | "
            "{message}"
        ),
        rotation="00:00",
        retention="30 days",
        compression="zip",
        backtrace=True,
        diagnose=True,
        encoding="utf-8",
        enqueue=True,
    )
    logger.info(f"Logs writing to: {_LOG_DIR.resolve()}")


# ── Excel reading ──────────────────────────────────────────────────────────────

def _read_pr_nos_from_excel(
    excel_path: Path,
    column: str | None,
    sheet: str | None,
) -> list[str]:
    """Read PURCHASE_REQ_NO values from an Excel file.

    Parameters
    ----------
    excel_path:
        Path to the .xlsx / .xls file.
    column:
        Explicit column header to read.  If None, auto-detects from known names
        or falls back to the first column.
    sheet:
        Sheet name.  If None, uses the first sheet.

    Returns
    -------
    list[str]
        Deduplicated, non-empty PR numbers in the order they appear.
    """
    try:
        import openpyxl
    except ImportError:
        logger.critical("openpyxl is not installed — run: pip install openpyxl")
        sys.exit(1)

    if not excel_path.exists():
        logger.critical(f"Excel file not found: {excel_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    if sheet:
        if sheet not in wb.sheetnames:
            logger.critical(
                f"Sheet {sheet!r} not found in {excel_path.name}. "
                f"Available: {wb.sheetnames}"
            )
            sys.exit(1)
        ws = wb[sheet]
    else:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        logger.critical(f"Excel sheet is empty: {excel_path}")
        sys.exit(1)

    # First row is the header
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]

    # Determine which column index to use
    col_idx: int | None = None

    if column:
        # Explicit column name — match case-insensitively
        col_lower = column.strip().lower()
        for i, h in enumerate(headers):
            if h.lower() == col_lower:
                col_idx = i
                break
        if col_idx is None:
            logger.critical(
                f"Column {column!r} not found in sheet. "
                f"Available columns: {headers}"
            )
            sys.exit(1)
    else:
        # Auto-detect from known names
        for i, h in enumerate(headers):
            if h.lower() in _AUTO_DETECT_COLUMNS:
                col_idx = i
                logger.info(f"Auto-detected PR column: {h!r} (index {i})")
                break

        if col_idx is None:
            # Fall back to the first column
            col_idx = 0
            logger.info(
                f"No known PR column found — using first column: {headers[0]!r}"
            )

    # Extract values, skip empty/None
    seen: set[str] = set()
    pr_nos: list[str] = []
    for row in rows[1:]:
        if col_idx >= len(row):
            continue
        val = row[col_idx]
        if val is None:
            continue
        pr_no = str(val).strip()
        if pr_no and pr_no not in seen:
            seen.add(pr_no)
            pr_nos.append(pr_no)

    wb.close()
    return pr_nos


# ── DB helpers ────────────────────────────────────────────────────────────────

_FETCH_STATUS_SQL = (
    f"SELECT [PURCHASEFINALAPPROVALSTATUS] "
    f"FROM {AzureTables.PURCHASE_REQ_MST} "
    f"WHERE [PURCHASE_REQ_NO] = ?"
)


def _get_pr_approval_status(pr_no: str, conn_str: str) -> str | None:
    """Return PURCHASEFINALAPPROVALSTATUS for the PR, or None if not found."""
    try:
        conn = connect_with_retry(conn_str, autocommit=True)
        try:
            cursor = conn.cursor()
            cursor.execute(_FETCH_STATUS_SQL, pr_no)
            row = cursor.fetchone()
            if row is None:
                return None
            return str(row[0]).strip() if row[0] is not None else ""
        finally:
            conn.close()
    except Exception as exc:
        logger.warning(f"Could not fetch status for PR={pr_no!r} from purchase_req_mst: {exc}")
        return None


# ── DB smoke-test ──────────────────────────────────────────────────────────────

def _verify_connections(config: AppConfig) -> None:
    checks = [
        ("Azure SQL",   config.get_azure_conn_str()),
        ("On-prem SQL", config.get_ras_conn_str()),
    ]
    failed = False
    for label, conn_str in checks:
        try:
            conn = connect_with_retry(conn_str, autocommit=True)
            conn.cursor().execute("SELECT 1")
            conn.close()
            logger.info(f"DB connection OK: {label}")
        except Exception as exc:
            logger.critical(f"DB connection FAILED: {label} — {exc}")
            failed = True
    if failed:
        logger.critical("One or more DB connections failed — aborting")
        sys.exit(1)


# ── CLI ────────────────────────────────────────────────────────────────────────

def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="python run_pipeline_from_excel.py",
        description=(
            "Run the full pipeline (INGESTION → COMPLETE, all 8 stages) for PRs "
            "listed in an Excel file.  Each PR is reprocessed from scratch."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--excel",
        required=True,
        metavar="FILE",
        help="Path to the Excel file (.xlsx) containing purchase requisition numbers.",
    )
    parser.add_argument(
        "--column",
        default=None,
        metavar="COLUMN_NAME",
        help=(
            "Header name of the column that contains PURCHASE_REQ_NO values. "
            "If omitted, auto-detects from common names or uses the first column."
        ),
    )
    parser.add_argument(
        "--sheet",
        default=None,
        metavar="SHEET_NAME",
        help="Sheet name to read (default: first sheet).",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=None,
        metavar="N",
        help=(
            "Number of PRs to process in parallel. "
            "Defaults to EXCEL_PIPELINE_WORKERS from .env (default 3). "
            "Each worker opens its own DB connections."
        ),
    )
    return parser


def main() -> None:
    _configure_logging()
    args = _build_parser().parse_args()

    # ── Load config ────────────────────────────────────────────────────────
    try:
        config = AppConfig()
    except EnvironmentError as exc:
        logger.critical(f"Configuration error: {exc}")
        sys.exit(1)

    _verify_connections(config)

    # ── Read PR numbers from Excel ─────────────────────────────────────────
    excel_path = Path(args.excel).resolve()
    pr_nos = _read_pr_nos_from_excel(excel_path, args.column, args.sheet)

    if not pr_nos:
        logger.warning("No purchase requisition numbers found in the Excel file — nothing to do")
        sys.exit(0)

    logger.info(
        f"Read {len(pr_nos)} unique PR number(s) from {excel_path.name} — "
        f"running full pipeline for each"
    )

    # ── Run pipeline for each PR ───────────────────────────────────────────
    orchestrator       = PipelineOrchestrator(config)
    azure_conn_str     = config.get_azure_conn_str()
    allowed_statuses   = config.EXCEL_ALLOWED_APPROVAL_STATUSES  # list[str], may be empty
    workers            = max(1, args.workers if args.workers is not None else config.EXCEL_PIPELINE_WORKERS)
    total              = len(pr_nos)

    if allowed_statuses:
        logger.info(
            f"Approval-status filter active — only processing PRs with "
            f"PURCHASEFINALAPPROVALSTATUS in: {allowed_statuses}"
        )
    else:
        logger.info("No approval-status filter configured — processing all PRs from Excel")

    succeeded = 0
    failed    = 0
    skipped   = 0

    def _process_one(pr_no: str, idx: int) -> str:
        """Run one PR through the full pipeline. Returns 'ok', 'failed', or 'skipped'."""
        status = _get_pr_approval_status(pr_no, azure_conn_str)

        if status is None:
            logger.warning(
                f"[{idx}/{total}] PR={pr_no!r} not found in purchase_req_mst — skipping"
            )
            return "skipped"

        if allowed_statuses and status.upper() not in allowed_statuses:
            logger.warning(
                f"[{idx}/{total}] PR={pr_no!r} skipped — "
                f"PURCHASEFINALAPPROVALSTATUS={status!r} not in allowed list"
            )
            return "skipped"

        logger.info(f"[{idx}/{total}] Starting full pipeline for PR={pr_no!r}")
        try:
            result = orchestrator.run_single(pr_no)
            if result.succeeded:
                logger.info(f"[{idx}/{total}] PR={pr_no!r} — SUCCESS")
                return "ok"
            else:
                stage = result.failed_stage
                logger.error(
                    f"[{idx}/{total}] PR={pr_no!r} — FAILED at "
                    f"stage={stage.stage_name!r}: {stage.error}"
                )
                return "failed"
        except Exception as exc:
            logger.opt(exception=True).error(
                f"[{idx}/{total}] PR={pr_no!r} — unexpected error: {exc}"
            )
            return "failed"

    if workers == 1:
        for idx, pr_no in enumerate(pr_nos, 1):
            outcome = _process_one(pr_no, idx)
            if outcome == "ok":
                succeeded += 1
            elif outcome == "failed":
                failed += 1
            else:
                skipped += 1
    else:
        logger.info(f"Running with {workers} parallel workers")
        counter_lock = threading.Lock()
        completed    = 0

        with ThreadPoolExecutor(max_workers=workers) as executor:
            future_to_pr = {
                executor.submit(_process_one, pr_no, i): pr_no
                for i, pr_no in enumerate(pr_nos, 1)
            }
            for future in as_completed(future_to_pr):
                pr_no = future_to_pr[future]
                with counter_lock:
                    completed += 1

                try:
                    outcome = future.result()
                except Exception as exc:
                    logger.opt(exception=True).error(
                        f"PR={pr_no!r} — unexpected thread error: {exc}"
                    )
                    outcome = "failed"

                if outcome == "ok":
                    with counter_lock:
                        succeeded += 1
                elif outcome == "failed":
                    with counter_lock:
                        failed += 1
                else:
                    with counter_lock:
                        skipped += 1

    # ── Summary ────────────────────────────────────────────────────────────
    logger.info("=" * 60)
    logger.info("Excel pipeline run summary")
    logger.info(f"  Excel file : {excel_path.name}")
    logger.info(f"  Workers    : {workers}")
    logger.info(f"  Total PRs  : {total}")
    logger.info(f"  Succeeded  : {succeeded}")
    logger.info(f"  Failed     : {failed}")
    logger.info(f"  Skipped    : {skipped}  (not found in purchase_req_mst, or status not in allowed list)")
    logger.info("=" * 60)

    sys.exit(1 if failed else 0)


if __name__ == "__main__":
    main()
